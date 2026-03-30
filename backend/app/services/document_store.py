import json, os, uuid, shutil
from datetime import datetime
from pathlib import Path
from pydantic import BaseModel

UPLOAD_DIR = Path(__file__).parent.parent / "data" / "uploads"
INDEX_FILE = Path(__file__).parent.parent / "data" / "doc_index.json"
STATES_FILE = Path(__file__).parent.parent / "data" / "doc_states.json"
KB_FILE = Path(__file__).parent.parent / "data" / "knowledge_base.json"
MAX_TEXT_LEN = 50_000

class DocumentMeta(BaseModel):
    id: str
    filename: str
    original_name: str
    file_type: str
    scope: str  # "global" | "session"
    session_id: str | None = None
    uploaded_at: str
    size_bytes: int
    status: str = "ready"
    text_preview: str = ""  # first 200 chars
    is_active: bool = True  # if False — excluded from context

class DocumentStore:
    def __init__(self):
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        self._docs: dict[str, DocumentMeta] = {}
        self._texts: dict[str, str] = {}  # id -> full extracted text
        self._load_index()
        self._register_default_kb()
        self._apply_saved_states()

    def _load_active_states(self) -> dict[str, bool]:
        """Load persisted active/inactive states for all documents."""
        if STATES_FILE.exists():
            try:
                return json.loads(STATES_FILE.read_text("utf-8"))
            except Exception:
                pass
        return {}

    def _save_active_states(self):
        """Persist active/inactive states for all documents."""
        states = {doc_id: doc.is_active for doc_id, doc in self._docs.items()}
        STATES_FILE.write_text(json.dumps(states, ensure_ascii=False, indent=2), "utf-8")

    def _apply_saved_states(self):
        """Apply persisted is_active states after all docs are registered."""
        states = self._load_active_states()
        for doc_id, is_active in states.items():
            if doc_id in self._docs:
                self._docs[doc_id].is_active = is_active

    def _load_index(self):
        if INDEX_FILE.exists():
            try:
                data = json.loads(INDEX_FILE.read_text("utf-8"))
                for d in data:
                    meta = DocumentMeta(**d)
                    # verify file still exists
                    fpath = UPLOAD_DIR / meta.filename
                    if fpath.exists():
                        self._docs[meta.id] = meta
                        self._texts[meta.id] = self._extract_text(fpath, meta.file_type)
            except Exception:
                pass

    def _register_default_kb(self):
        if "kb_default" not in self._docs and KB_FILE.exists():
            stat = KB_FILE.stat()
            self._docs["kb_default"] = DocumentMeta(
                id="kb_default",
                filename="knowledge_base.json",
                original_name="knowledge_base.json (АФК Система)",
                file_type="json",
                scope="global",
                uploaded_at=datetime.fromtimestamp(stat.st_mtime).isoformat(),
                size_bytes=stat.st_size,
                status="ready",
                text_preview="База знаний АФК Система: 23 компании, финансы, события..."
            )
            text = KB_FILE.read_text("utf-8")
            self._texts["kb_default"] = text[:MAX_TEXT_LEN]
            try:
                from app.services.vector_store import vector_store
                vector_store.index_document("kb_default", "knowledge_base.json", text[:MAX_TEXT_LEN])
            except Exception:
                pass
        self._register_demo_sources()

    def _register_demo_sources(self):
        """Register demo source documents with full realistic content from hardcoded data modules."""
        from app.data.project import PROJECT
        from app.data.market import MARKET_DATA
        from app.data.committee import COMMITTEE_MATERIALS

        mkt = MARKET_DATA["logistics"]
        cm = COMMITTEE_MATERIALS

        demo_docs = [
            # --- Отчетности портфельных компаний ---
            ("src_mts", "Годовой отчет МТС 2025", "pdf", 2_450_000,
             "Годовой отчет ПАО МТС за 2025 год",
             "ГОДОВОЙ ОТЧЕТ ПАО «МТС» ЗА 2025 ГОД\n\n"
             "КЛЮЧЕВЫЕ ФИНАНСОВЫЕ ПОКАЗАТЕЛИ\n"
             "Выручка: 807.2 млрд руб. (+14.7% г/г)\n"
             "EBITDA: 340.0 млрд руб.\n"
             "Маржа EBITDA: 42.1%\n"
             "Чистая прибыль: 32.8 млрд руб.\n"
             "Чистый долг: 544.0 млрд руб.\n"
             "Чистый долг/EBITDA: 1.6x\n\n"
             "ДИВИДЕНДНАЯ ПОЛИТИКА\n"
             "Рекомендованный дивиденд: 35 руб./акция\n"
             "Дивидендная доходность: ~12.5%\n"
             "Периодичность: 2 раза в год\n\n"
             "СТРАТЕГИЧЕСКИЕ ПРИОРИТЕТЫ\n"
             "1. Развитие экосистемы МТС (финтех, медиа, облачные сервисы)\n"
             "2. 5G — инвестиции в инфраструктуру\n"
             "3. Рост ARPU через пакетные предложения\n"
             "4. Международная экспансия (СНГ)\n\n"
             "СЕГМЕНТНАЯ СТРУКТУРА ВЫРУЧКИ\n"
             "Мобильная связь: 58%\n"
             "Фиксированная связь: 15%\n"
             "Финтех (МТС Банк): 12%\n"
             "Медиа и развлечения: 8%\n"
             "Облачные сервисы и IoT: 7%"),

            ("src_seg_report", "Консолидированная отчетность Segezha 2025", "pdf", 1_800_000,
             "Консолидированная финансовая отчетность Segezha Group",
             "КОНСОЛИДИРОВАННАЯ ФИНАНСОВАЯ ОТЧЕТНОСТЬ\nSEGEZHA GROUP ЗА 2025 ГОД\n\n"
             "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ\n"
             "Выручка: 101.9 млрд руб. (+11.3% г/г)\n"
             "EBITDA: 7.1 млрд руб.\n"
             "Маржа EBITDA: 7.0%\n"
             "Чистый убыток: -28.0 млрд руб.\n"
             "Чистый долг: 102.2 млрд руб.\n"
             "Чистый долг/EBITDA: 14.4x (КРИТИЧЕСКИЙ УРОВЕНЬ)\n\n"
             "ДОЛГОВАЯ НАГРУЗКА\n"
             "Синдицированный кредит: 65 млрд руб.\n"
             "Облигации: 25 млрд руб.\n"
             "Прочие кредиты: 12.2 млрд руб.\n"
             "Ковенант Чистый долг/EBITDA: не более 4.5x (НАРУШЕН)\n"
             "Статус: ведутся переговоры о реструктуризации с банками\n\n"
             "ОПЕРАЦИОННЫЕ ПОКАЗАТЕЛИ\n"
             "Производство пиломатериалов: 4.2 млн куб.м\n"
             "Производство фанеры: 320 тыс. куб.м\n"
             "Производство бумаги: 870 тыс. тонн\n\n"
             "РЫНОЧНАЯ КОНЪЮНКТУРА\n"
             "Цены на пиломатериалы восстанавливаются после минимумов 2023-2024\n"
             "Экспорт ограничен санкциями — переориентация на Китай и Среднюю Азию"),

            ("src_seg_q4", "Квартальный отчет Segezha Q4 2025", "pdf", 890_000,
             "Квартальный отчет Segezha Group за Q4 2025",
             "КВАРТАЛЬНЫЙ ОТЧЕТ SEGEZHA GROUP\nQ4 2025\n\n"
             "Выручка Q4: 28.5 млрд руб. (+15% кв/кв)\n"
             "EBITDA Q4: 2.1 млрд руб.\n"
             "Чистый долг/EBITDA (LTM): 14.4x\n\n"
             "ДИНАМИКА ДОЛГОВОЙ НАГРУЗКИ\n"
             "Q1 2025: 15.8x\n"
             "Q2 2025: 15.1x\n"
             "Q3 2025: 14.9x\n"
             "Q4 2025: 14.4x (улучшение, но далеко от ковенанта 4.5x)\n\n"
             "ПРОГНОЗ МЕНЕДЖМЕНТА\n"
             "Целевой Чистый долг/EBITDA к 2027: 5.0x\n"
             "Необходимый рост EBITDA: до 20+ млрд руб.\n"
             "Допэмиссия: рассматривается как опция"),

            ("src_etalon", "МСФО Эталон 2025", "pdf", 1_200_000,
             "Отчетность по МСФО ГК Эталон за 2025 год",
             "ОТЧЕТНОСТЬ ПО МСФО\nГК ЭТАЛОН ЗА 2025 ГОД\n\n"
             "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ\n"
             "Выручка: 115.3 млрд руб. (+2.1% г/г)\n"
             "EBITDA: 20.8 млрд руб.\n"
             "Маржа EBITDA: 18.0%\n"
             "Чистая прибыль: 8.2 млрд руб.\n"
             "Чистый долг: 51.9 млрд руб.\n"
             "Чистый долг/EBITDA: 2.5x\n\n"
             "ОПЕРАЦИОННЫЕ ПОКАЗАТЕЛИ\n"
             "Продажи: 1.2 млн кв.м\n"
             "Средняя цена реализации: 96 000 руб./кв.м\n"
             "Портфель проектов: 8.5 млн кв.м\n\n"
             "СОБЫТИЯ\n"
             "Допэмиссия в пользу АФК Система (увеличение доли до 51%)\n"
             "Географическая экспансия: Краснодарский край, Тюмень"),

            ("src_step", "Годовой отчет СТЕПЬ 2025", "pdf", 1_500_000,
             "Годовой отчет ГК СТЕПЬ за 2025 год",
             "ГОДОВОЙ ОТЧЕТ ГК «СТЕПЬ» ЗА 2025 ГОД\n\n"
             "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ\n"
             "Выручка: 68.0 млрд руб. (+18.0% г/г)\n"
             "EBITDA: 10.2 млрд руб.\n"
             "Маржа EBITDA: 15.0%\n"
             "Чистая прибыль: 5.0 млрд руб.\n"
             "Чистый долг/EBITDA: 1.3x\n\n"
             "ОПЕРАЦИОННЫЕ ПОКАЗАТЕЛИ\n"
             "Земельный банк: 600 000 га\n"
             "Валовой сбор зерновых: 3.2 млн тонн\n"
             "Урожайность: 55 ц/га (выше среднего по РФ)\n\n"
             "IPO ПЛАНЫ\n"
             "Подготовка к IPO в 2026-2027 году\n"
             "Ожидаемая оценка: 80-100 млрд руб."),

            ("src_binno", "Отчетность Биннофарм 2025", "pdf", 950_000,
             "Финансовая отчетность Биннофарм Групп",
             "ФИНАНСОВАЯ ОТЧЕТНОСТЬ\nБИННОФАРМ ГРУПП ЗА 2025 ГОД\n\n"
             "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ\n"
             "Выручка: 45.0 млрд руб. (+16.0% г/г)\n"
             "EBITDA: 9.9 млрд руб.\n"
             "Маржа EBITDA: 22.0%\n"
             "Чистая прибыль: 5.5 млрд руб.\n"
             "Чистый долг/EBITDA: 0.8x\n\n"
             "ПРОДУКТОВЫЙ ПОРТФЕЛЬ\n"
             "Дженерики: 65% выручки\n"
             "Оригинальные препараты: 20%\n"
             "Биопрепараты: 15%\n\n"
             "IPO ПЛАНЫ\n"
             "IPO запланировано на 2026 год\n"
             "Ожидаемая оценка: 60-80 млрд руб.\n"
             "Привлечение: 15-20 млрд руб."),

            # --- Юридические и корпоративные документы ---
            ("src_seg_credit", "Кредитное соглашение Segezha (синдицированный кредит)", "pdf", 340_000,
             "Синдицированное кредитное соглашение Segezha",
             "СИНДИЦИРОВАННОЕ КРЕДИТНОЕ СОГЛАШЕНИЕ\nSEGEZHA GROUP\n\n"
             "Дата: 15 марта 2023 г.\nСумма: 65 000 000 000 руб.\nСрок: 5 лет\n\n"
             "8. ФИНАНСОВЫЕ КОВЕНАНТЫ\n\n"
             "8.1. Заемщик обязуется поддерживать следующие финансовые показатели:\n\n"
             "8.2. Чистый долг / EBITDA\n"
             "Значение: не более 4.5x (на каждую отчетную дату)\n"
             "Метод расчета: LTM EBITDA по МСФО\n"
             "Текущее значение (Q4 2025): 14.4x — НАРУШЕНИЕ КОВЕНАНТА\n\n"
             "8.3. Коэффициент покрытия процентов (ICR)\n"
             "Значение: не менее 2.0x\n"
             "Текущее значение: 0.8x — НАРУШЕНИЕ КОВЕНАНТА\n\n"
             "9. ПОСЛЕДСТВИЯ НАРУШЕНИЯ\n"
             "9.1. Банки вправе потребовать досрочного погашения\n"
             "9.2. Применяется штрафная ставка +2% годовых\n"
             "9.3. Ведутся переговоры о waiver и реструктуризации"),

            ("src_div_mts", "Дивидендная политика МТС", "pdf", 120_000,
             "Дивидендная политика ПАО МТС",
             "ДИВИДЕНДНАЯ ПОЛИТИКА ПАО «МТС»\n\n"
             "Утверждена Советом директоров 25.02.2024\n\n"
             "1. ЦЕЛЕВОЙ УРОВЕНЬ ВЫПЛАТ\n"
             "Не менее 35 руб. на акцию в год\n"
             "Целевой payout ratio: 50-70% от чистой прибыли по МСФО\n\n"
             "2. ПЕРИОДИЧНОСТЬ\n"
             "Выплата дивидендов 2 раза в год (промежуточные + финальные)\n\n"
             "3. ИСТОРИЯ ДИВИДЕНДОВ\n"
             "2023: 34.29 руб./акция\n"
             "2024: 35.00 руб./акция\n"
             "2025E: 35.00 руб./акция (рекомендация СД)\n\n"
             "4. ДИВИДЕНДНАЯ ДОХОДНОСТЬ\n"
             "При текущей цене акции ~280 руб.: доходность ~12.5%"),

            ("src_seg_div", "Решение СД Segezha о приостановке дивидендов", "pdf", 85_000,
             "Протокол решения Совета директоров Segezha Group",
             "ПРОТОКОЛ ЗАСЕДАНИЯ СОВЕТА ДИРЕКТОРОВ\nSEGEZHA GROUP\n\n"
             "Дата: 18 ноября 2025 г.\n\n"
             "ПОВЕСТКА ДНЯ\n"
             "Вопрос 3: О приостановке выплаты дивидендов\n\n"
             "РЕШЕНИЕ (единогласно):\n"
             "Приостановить выплату дивидендов на неопределенный срок\n"
             "в связи с высоким уровнем долговой нагрузки (Чистый долг/EBITDA 14.4x)\n"
             "и необходимостью направления свободного денежного потока на обслуживание долга.\n\n"
             "Возобновление дивидендных выплат возможно при достижении\n"
             "Чистый долг/EBITDA ≤ 3.0x"),

            # --- Инвестиционный проект: Логистический хаб ---
            ("src_project", "Презентация проекта (Логистический хаб Подмосковье)", "pdf", 3_200_000,
             "Инвестиционная презентация проекта",
             f"ИНВЕСТИЦИОННАЯ ПРЕЗЕНТАЦИЯ\nЛОГИСТИЧЕСКИЙ ХАБ «ПОДМОСКОВЬЕ»\n\n"
             f"1. ОПИСАНИЕ ПРОЕКТА\n"
             f"Тип: {PROJECT['type']}\n"
             f"Площадь: {PROJECT['area_sqm']:,} кв.м\n"
             f"Локация: Московская область, Домодедовский район\n"
             f"Инвестиции: {PROJECT['investment_bln']} млрд руб.\n"
             f"Горизонт: {PROJECT['horizon_years']} лет\n\n"
             f"2. ФИНАНСОВЫЕ ПАРАМЕТРЫ\n"
             f"Ставка аренды: {PROJECT['rent_rate_sqm_year']:,} руб./кв.м/год\n"
             f"Целевая заполняемость: {PROJECT['occupancy_rate']*100:.0f}%\n"
             f"OPEX: {PROJECT['opex_pct']*100:.0f}% от выручки\n"
             f"Стоимость строительства: {PROJECT['construction_cost_sqm']:,} руб./кв.м\n"
             f"Ставка дисконтирования: {PROJECT['discount_rate']*100:.0f}%\n"
             f"Terminal cap rate: {PROJECT['terminal_cap_rate']*100:.0f}%\n\n"
             f"3. РЕЗУЛЬТАТЫ ОЦЕНКИ\n"
             f"NPV: {PROJECT['npv']} млрд руб.\n"
             f"IRR: {PROJECT['irr']*100:.1f}%\n"
             f"Срок окупаемости: {PROJECT['payback_years']} лет\n\n"
             f"4. КОНКУРЕНТНЫЕ ПРЕИМУЩЕСТВА\n"
             f"— Минимальная конкуренция в радиусе 30 км\n"
             f"— Близость к ЦКАД и федеральным трассам\n"
             f"— Дефицит складов класса А (вакансия 3.2%)\n"
             f"— Якорный арендатор: предварительное соглашение на 30% площадей"),

            ("src_finmodel", "Финансовая модель проекта (XLSX)", "xlsx", 450_000,
             "Финансовая модель логистического хаба",
             f"ФИНАНСОВАЯ МОДЕЛЬ\nЛОГИСТИЧЕСКИЙ ХАБ «ПОДМОСКОВЬЕ»\n\n"
             f"=== ВКЛАДКА: ДОПУЩЕНИЯ ===\n"
             f"Ячейка C5: Площадь = {PROJECT['area_sqm']:,} кв.м\n"
             f"Ячейка C8: Ставка аренды = {PROJECT['rent_rate_sqm_year']:,} руб./кв.м/год\n"
             f"Ячейка C12: Заполняемость = {PROJECT['occupancy_rate']*100:.0f}%\n"
             f"Ячейка C15: OPEX = {PROJECT['opex_pct']*100:.0f}% от выручки\n"
             f"Ячейка C18: WACC = {PROJECT['discount_rate']*100:.0f}%\n\n"
             f"=== ВКЛАДКА: CF (Cash Flow) ===\n"
             f"Ячейка B3: CAPEX = {PROJECT['investment_bln']} млрд руб.\n\n"
             "Год\tCash Flow (млн руб.)\n" +
             "\n".join(f"  {i}\t{cf:,.0f}" for i, cf in enumerate(PROJECT['cash_flows'])) +
             f"\n\n=== ВКЛАДКА: РЕЗУЛЬТАТЫ ===\n"
             f"NPV = {PROJECT['npv']} млрд руб.\n"
             f"IRR = {PROJECT['irr']*100:.1f}%\n"
             f"Payback = {PROJECT['payback_years']} лет\n\n"
             f"=== ВКЛАДКА: ЧУВСТВИТЕЛЬНОСТЬ ===\n"
             f"NPV (млрд руб.) при разных заполняемости и ставке аренды:\n"
             f"Заполняемость →\t" + "\t".join(f"{r*100:.0f}%" for r in PROJECT['sensitivity']['occupancy_rates']) + "\n" +
             "\n".join(
                 f"Аренда {rent}\t" + "\t".join(f"{v:.1f}" for v in row)
                 for rent, row in zip(PROJECT['sensitivity']['rent_rates'], PROJECT['sensitivity']['npv_matrix'])
             )),

            ("src_appraiser", "Отчет независимого оценщика", "pdf", 1_100_000,
             "Независимая оценка стоимости проекта",
             f"ОТЧЕТ НЕЗАВИСИМОГО ОЦЕНЩИКА\nПроект: Логистический хаб «Подмосковье»\n"
             f"Оценщик: ООО «Профессиональная Оценка»\nДата: 01.12.2025\n\n"
             f"1. РЫНОЧНАЯ СТОИМОСТЬ ПРОЕКТА\n"
             f"Справедливая стоимость (DCF): {cm['documents'][2]['key_claims']['market_value']} млрд руб.\n"
             f"Cap rate: {cm['documents'][2]['key_claims']['cap_rate']*100:.0f}%\n\n"
             f"2. РЫНОЧНЫЕ БЕНЧМАРКИ\n"
             f"Средняя заполняемость по рынку: {cm['documents'][2]['key_claims']['occupancy_market_avg']*100:.0f}%\n"
             f"Средняя ставка аренды (класс А): {cm['documents'][2]['key_claims']['rent_market_avg']:,} руб./кв.м/год\n"
             f"Стоимость строительства: {cm['documents'][2]['key_claims']['construction_cost_market']}\n\n"
             f"3. КОНКУРЕНТНАЯ СРЕДА (стр. 31)\n"
             f"В радиусе 30 км выявлено: {cm['documents'][2]['key_claims']['competitors_in_radius']}\n"
             f"— PNK Park Домодедово (125 000 кв.м, заполняемость 91%)\n"
             f"— Логопарк Север-2 (80 000 кв.м, заполняемость 88%)\n"
             f"— А-Терминал (45 000 кв.м, заполняемость 94%)\n\n"
             f"4. ВЫВОДЫ\n"
             f"Проект экономически обоснован при заполняемости не ниже 90%.\n"
             f"Заявленная заполняемость 95-97% представляется оптимистичной."),

            ("src_mgmt", "Презентация менеджмента", "pdf", 2_800_000,
             "Презентация управляющей команды проекта",
             f"ПРЕЗЕНТАЦИЯ МЕНЕДЖМЕНТА\nЛогистический хаб «Подмосковье»\n\n"
             f"КЛЮЧЕВЫЕ ЗАЯВЛЕНИЯ МЕНЕДЖМЕНТА\n\n"
             f"стр. 8: Целевая заполняемость: {cm['documents'][0]['key_claims']['occupancy']*100:.0f}%\n"
             f"стр. 12: Объем инвестиций: {cm['documents'][0]['key_claims']['capex']} млрд руб.\n"
             f"стр. 14: Рост выручки: {cm['documents'][0]['key_claims']['revenue_growth']}\n"
             f"стр. 16: IRR проекта: {cm['documents'][0]['key_claims']['irr']*100:.0f}%\n"
             f"стр. 15: Рыночная позиция: {cm['documents'][0]['key_claims']['market_position']}\n"
             f"стр. 15: Конкуренция: {cm['documents'][0]['key_claims']['competitors']}\n\n"
             f"КОМАНДА ПРОЕКТА\n"
             f"CEO: Иванов А.В. (15 лет в девелопменте)\n"
             f"CFO: Петрова М.С. (ex-JLL Russia)\n"
             f"COO: Сидоров К.Н. (ex-PNK Group)"),

            ("src_legal_dd", "Юридический due diligence", "pdf", 450_000,
             "Юридическое заключение по проекту",
             f"ЮРИДИЧЕСКИЙ DUE DILIGENCE\nПроект: Логистический хаб «Подмосковье»\n"
             f"Юридическая фирма: Пепеляев Групп\nДата: 10.12.2025\n\n"
             f"1. ЗЕМЕЛЬНЫЙ УЧАСТОК\n"
             f"Статус: {cm['documents'][3]['key_claims']['land_status']}\n"
             f"Площадь: 12.5 га\n"
             f"Категория: земли промышленности\n\n"
             f"2. РАЗРЕШЕНИЯ И СОГЛАСОВАНИЯ\n"
             f"{cm['documents'][3]['key_claims']['permits']}\n\n"
             f"3. ОБРЕМЕНЕНИЯ\n"
             f"Сервитут: {cm['documents'][3]['key_claims']['encumbrances']}\n\n"
             f"4. ВЫЯВЛЕННЫЕ РИСКИ\n" +
             "\n".join(f"— {r}" for r in cm['documents'][3]['key_claims']['risks']) +
             f"\n\n5. ОБЩЕЕ ЗАКЛЮЧЕНИЕ\n"
             f"Правовой статус объекта позволяет реализацию проекта.\n"
             f"Рекомендации: урегулировать судебный спор до начала строительства."),

            # --- Рыночная аналитика ---
            ("src_logistics", "Аналитика рынка логистики РФ 2025", "pdf", 1_600_000,
             "Аналитический обзор рынка логистики России",
             f"АНАЛИТИЧЕСКИЙ ОБЗОР\nРЫНОК ЛОГИСТИКИ И СКЛАДСКОЙ НЕДВИЖИМОСТИ РОССИИ 2025\n\n"
             f"1. РАЗМЕР РЫНКА\n"
             f"Объем: {mkt['size_2025_trln']} трлн руб. (2025)\n"
             f"CAGR (2021-2025): {mkt['cagr_pct']}%\n\n"
             f"Динамика:\n" +
             "\n".join(f"  {h['year']}: {h['size']} трлн руб." for h in mkt['size_history']) +
             f"\n\n2. СТРУКТУРА ПО СЕГМЕНТАМ\n" +
             "\n".join(f"  {s['name']}: {s['share']}% (маржа {s['margin']})" for s in mkt['segments']) +
             f"\n\n3. КЛЮЧЕВЫЕ ИГРОКИ\n" +
             "\n".join(f"  {p['name']}: доля {p['share']}%, выручка {p['revenue']} млрд руб." for p in mkt['players']) +
             f"\n\n4. СКЛАДСКОЙ РЫНОК\n"
             f"Общий объем: {mkt['warehouse_metrics']['total_stock_mln_sqm']} млн кв.м\n"
             f"Вакансия: {mkt['warehouse_metrics']['vacancy_rate_pct']}%\n"
             f"Средняя ставка (класс А): {mkt['warehouse_metrics']['avg_rent_class_a']:,} руб./кв.м/год\n"
             f"Новое предложение 2025: {mkt['warehouse_metrics']['new_supply_2025_mln_sqm']} млн кв.м\n"
             f"Доля Москвы: {mkt['warehouse_metrics']['moscow_share_pct']}%\n\n"
             f"5. ТРЕНДЫ\n" +
             "\n".join(f"  — {t}" for t in mkt['trends'])),

            ("src_jll", "JLL Russia Warehouse Market Q4 2025", "pdf", 2_100_000,
             "Обзор складского рынка России от JLL",
             f"JLL RUSSIA\nWAREHOUSE MARKET OVERVIEW Q4 2025\n\n"
             f"KEY METRICS\n"
             f"Total stock: {mkt['warehouse_metrics']['total_stock_mln_sqm']} mln sqm\n"
             f"Vacancy rate: {mkt['warehouse_metrics']['vacancy_rate_pct']}% (historical low)\n"
             f"Class A asking rent: {mkt['warehouse_metrics']['avg_rent_class_a']:,} RUB/sqm/year (+18% YoY)\n"
             f"New supply 2025: {mkt['warehouse_metrics']['new_supply_2025_mln_sqm']} mln sqm\n\n"
             f"M&A DEALS\n" +
             "\n".join(f"  {d['year']}: {d['buyer']} → {d['target']} ({d['value']})" for d in mkt['ma_deals']) +
             f"\n\nMARKET RISKS\n" +
             "\n".join(f"  — {r}" for r in mkt['risks'])),

            ("src_bloomberg", "Данные Bloomberg Terminal", "xlsx", 780_000,
             "Рыночные мультипликаторы и котировки",
             "BLOOMBERG TERMINAL DATA EXPORT\nDate: 2025-12-20\n\n"
             "=== SECTOR MULTIPLES ===\n"
             "Sector\tEV/EBITDA\tP/E\tDiv Yield\n"
             "Telecom (RU)\t5.2x\t8.5x\t10-13%\n"
             "Real Estate (RU)\t7.8x\t12.0x\t2-5%\n"
             "Forestry (RU)\t8.5x\tn/a\t0%\n"
             "Pharma (RU)\t10.2x\t15.0x\t1-3%\n"
             "Agro (RU)\t6.0x\t9.0x\t3-5%\n"
             "Logistics (RU)\t8.0x\t14.0x\t1-2%\n\n"
             "=== AFK SISTEMA PORTFOLIO ===\n"
             "Ticker\tPrice\tMkt Cap (B)\tEV/EBITDA\n"
             "MTSS RX\t280 RUB\t561 B\t4.9x\n"
             "SGZH RX\t3.2 RUB\t50 B\tn/m (убыток)\n"
             "ETLN RX\t65 RUB\t78 B\t6.8x\n"
             "OZON RX\t3200 RUB\t640 B\t12.5x\n"),

            ("src_industry", "Отраслевые обзоры аналитиков", "pdf", 3_500_000,
             "Сводка отраслевых обзоров",
             "СВОДКА ОТРАСЛЕВЫХ ОБЗОРОВ\nМарт 2026\n\n"
             "ТЕЛЕКОМ\n"
             "МТС: целевая цена 320 руб. (upside 14%), рекомендация ПОКУПАТЬ\n"
             "Драйверы: 5G, финтех экосистема, стабильные дивиденды\n\n"
             "ЛЕСОПРОМ\n"
             "Segezha: целевая цена 2.5 руб. (downside 22%), рекомендация ПРОДАВАТЬ\n"
             "Риски: долговая нагрузка, ковенанты, реструктуризация\n\n"
             "НЕДВИЖИМОСТЬ\n"
             "Эталон: целевая цена 80 руб. (upside 23%), рекомендация ПОКУПАТЬ\n"
             "Драйверы: ипотечные программы, региональная экспансия\n\n"
             "ФАРМА\n"
             "Биннофарм: pre-IPO оценка 60-80 млрд руб.\n"
             "Драйверы: импортозамещение, рост потребления\n\n"
             "АГРО\n"
             "СТЕПЬ: pre-IPO оценка 80-100 млрд руб.\n"
             "Драйверы: экспорт, высокая урожайность, рост цен"),

            # --- Материалы инвесткомитета ---
            ("src_committee", "Материалы инвесткомитета (полный комплект)", "pdf", 5_200_000,
             "Полный комплект материалов к заседанию",
             "МАТЕРИАЛЫ К ЗАСЕДАНИЮ ИНВЕСТИЦИОННОГО КОМИТЕТА\nАФК СИСТЕМА\n"
             "Дата заседания: 20.01.2026\n\n"
             "ПОВЕСТКА ДНЯ\n"
             "Вопрос 1: Рассмотрение инвестпроекта «Логистический хаб Подмосковье»\n\n"
             "КОМПЛЕКТ ДОКУМЕНТОВ:\n"
             "1. Презентация менеджмента (24 стр.)\n"
             "2. Финансовая модель (15 вкладок)\n"
             "3. Отчет независимого оценщика (45 стр.)\n"
             "4. Юридический due diligence (18 стр.)\n\n"
             "ВЫЯВЛЕННЫЕ ПРОТИВОРЕЧИЯ МЕЖДУ ДОКУМЕНТАМИ:\n\n" +
             "\n\n".join(
                 f"Противоречие {i+1}: {c['parameter']}\n"
                 f"  {c['doc1']}: {c['doc1_value']} ({c['doc1_page']})\n"
                 f"  {c['doc2']}: {c['doc2_value']} ({c['doc2_page']})\n" +
                 (f"  {c.get('doc3', 'Оценщик')}: {c.get('doc3_value', '')} ({c.get('doc3_page', '')})\n" if c.get('doc3_value') else "") +
                 f"  Критичность: {c['severity']}\n"
                 f"  Комментарий: {c['comment']}"
                 for i, c in enumerate(cm['contradictions'])
             ) +
             "\n\nКЛЮЧЕВЫЕ РИСКИ:\n\n" +
             "\n\n".join(
                 f"Риск: {r['name']}\n"
                 f"  Критичность: {r['severity']}, Вероятность: {r['probability']}\n"
                 f"  Влияние: {r['impact']}\n"
                 f"  Источник: {r['source']}"
                 for r in cm['risks']
             )),
        ]

        for doc_id, name, ftype, size, preview, content in demo_docs:
            if doc_id not in self._docs:
                self._docs[doc_id] = DocumentMeta(
                    id=doc_id,
                    filename=f"{doc_id}.{ftype}",
                    original_name=name,
                    file_type=ftype,
                    scope="global",
                    uploaded_at="2025-12-15T10:00:00",
                    size_bytes=size,
                    status="ready",
                    text_preview=preview,
                )
                self._texts[doc_id] = content

        # Index all demo docs in vector store
        try:
            from app.services.vector_store import vector_store
            for doc_id, meta in self._docs.items():
                if doc_id != "kb_default":
                    text = self._texts.get(doc_id, "")
                    if text:
                        vector_store.index_document(doc_id, meta.original_name, text)
        except Exception:
            pass

    def _persist_index(self):
        # Only persist non-default docs
        docs = [d.model_dump() for d in self._docs.values() if d.id != "kb_default"]
        INDEX_FILE.write_text(json.dumps(docs, ensure_ascii=False, indent=2), "utf-8")

    def list_global(self) -> list[DocumentMeta]:
        return [d for d in self._docs.values() if d.scope == "global"]

    def list_session(self, session_id: str) -> list[DocumentMeta]:
        return [d for d in self._docs.values() if d.scope == "session" and d.session_id == session_id]

    def get(self, doc_id: str) -> DocumentMeta | None:
        return self._docs.get(doc_id)

    def get_text(self, doc_id: str) -> str:
        return self._texts.get(doc_id, "")

    def add(self, file_bytes: bytes, original_name: str, scope: str = "global", session_id: str | None = None) -> DocumentMeta:
        doc_id = str(uuid.uuid4())[:8]
        ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else "txt"
        filename = f"{doc_id}.{ext}"
        fpath = UPLOAD_DIR / filename
        fpath.write_bytes(file_bytes)

        text = self._extract_text(fpath, ext)
        meta = DocumentMeta(
            id=doc_id,
            filename=filename,
            original_name=original_name,
            file_type=ext,
            scope=scope,
            session_id=session_id,
            uploaded_at=datetime.now().isoformat(),
            size_bytes=len(file_bytes),
            status="ready",
            text_preview=text[:200]
        )
        self._docs[doc_id] = meta
        self._texts[doc_id] = text
        # Index in vector store
        try:
            from app.services.vector_store import vector_store
            vector_store.index_document(doc_id, original_name, text)
        except Exception:
            pass
        self._persist_index()
        # Invalidate response cache
        try:
            from app.services.skill_router import clear_response_cache
            clear_response_cache()
        except Exception:
            pass
        return meta

    def delete(self, doc_id: str) -> bool:
        meta = self._docs.pop(doc_id, None)
        if meta:
            self._texts.pop(doc_id, None)
            # Remove from vector store
            try:
                from app.services.vector_store import vector_store
                vector_store.remove_document(doc_id)
            except Exception:
                pass
            fpath = UPLOAD_DIR / meta.filename
            if fpath.exists():
                fpath.unlink()
            self._persist_index()
            # Invalidate response cache
            try:
                from app.services.skill_router import clear_response_cache
                clear_response_cache()
            except Exception:
                pass
            return True
        return False

    def toggle_active(self, doc_id: str) -> DocumentMeta | None:
        """Toggle is_active for a document. Returns updated meta or None if not found."""
        meta = self._docs.get(doc_id)
        if not meta:
            return None
        meta.is_active = not meta.is_active
        self._save_active_states()
        # Invalidate response cache
        try:
            from app.services.skill_router import clear_response_cache
            clear_response_cache()
        except Exception:
            pass
        return meta

    def get_context_for_skill(self, session_id: str | None = None) -> str:
        """Get concatenated text from all global docs + session docs"""
        parts = []
        for doc_id, meta in self._docs.items():
            if not meta.is_active:
                continue
            if meta.scope == "global" and doc_id != "kb_default":
                text = self._texts.get(doc_id, "")
                if text:
                    parts.append(f"--- {meta.original_name} ---\n{text}")
            elif meta.scope == "session" and session_id and meta.session_id == session_id:
                text = self._texts.get(doc_id, "")
                if text:
                    parts.append(f"--- {meta.original_name} ---\n{text}")
        return "\n\n".join(parts)

    def _extract_text(self, filepath: Path, file_type: str) -> str:
        try:
            if file_type in ("txt", "md"):
                return filepath.read_text("utf-8", errors="replace")[:MAX_TEXT_LEN]
            elif file_type == "json":
                data = json.loads(filepath.read_text("utf-8"))
                return json.dumps(data, ensure_ascii=False, indent=2)[:MAX_TEXT_LEN]
            elif file_type == "pdf":
                return self._extract_pdf(filepath)
            elif file_type == "docx":
                return self._extract_docx(filepath)
            elif file_type in ("xlsx", "xls"):
                return self._extract_xlsx(filepath)
            else:
                return filepath.read_text("utf-8", errors="replace")[:MAX_TEXT_LEN]
        except Exception as e:
            return f"[Ошибка извлечения текста: {e}]"

    def _extract_pdf(self, filepath: Path) -> str:
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages[:50]:  # max 50 pages
                    t = page.extract_text()
                    if t:
                        text_parts.append(t)
            return "\n\n".join(text_parts)[:MAX_TEXT_LEN]
        except ImportError:
            return "[pdfplumber not installed]"

    def _extract_docx(self, filepath: Path) -> str:
        try:
            from docx import Document
            doc = Document(str(filepath))
            text_parts = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(text_parts)[:MAX_TEXT_LEN]
        except ImportError:
            return "[python-docx not installed]"

    def _extract_xlsx(self, filepath: Path) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(filepath), read_only=True, data_only=True)
            text_parts = []
            for ws in wb.worksheets[:5]:  # max 5 sheets
                text_parts.append(f"Sheet: {ws.title}")
                for row in ws.iter_rows(max_row=200, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        text_parts.append("\t".join(cells))
            wb.close()
            return "\n".join(text_parts)[:MAX_TEXT_LEN]
        except ImportError:
            return "[openpyxl not installed]"


# Singleton
doc_store = DocumentStore()
